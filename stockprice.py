import pandas as pd
import yfinance as yf
from tabulate import tabulate
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import time

# --- CONFIGURATION ---
USD_TO_INR = 88.57
CSV_FILE_PATH = 'portfolio.csv'
EXCEL_OUTPUT_FILE = 'portfolio_analysis.xlsx'
MAX_RETRIES = 3
RETRY_DELAY = 2  # seconds between retries
BATCH_SIZE = 10  # number of tickers to fetch at once
BATCH_DELAY = 1  # seconds between batches

def read_portfolio_data(file_path: str) -> pd.DataFrame:
    """Reads the user's portfolio data from a CSV file."""
    try:
        df = pd.read_csv(file_path)
        required_cols = ['ticker', 'shares_held', 'avg_cost_usd']
        if not all(col in df.columns for col in required_cols):
            raise ValueError(f"CSV must contain columns: {required_cols}")
        print(f"✅ Loaded portfolio data for {len(df)} tickers.")
        return df
    except FileNotFoundError:
        print(f"❌ Error: The file '{file_path}' was not found.")
        with open(file_path, 'w') as f:
            f.write("ticker,shares_held,avg_cost_usd\nAAPL,10.0,175.50\nMSFT,5.5,350.00")
        print(f"Created a sample '{file_path}'. Please fill it with your data and run again.")
        exit()
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        exit()

def fetch_stock_data(tickers: List[str]) -> Dict[str, Dict[str, float]]:
    """Fetches current and previous close data using yfinance with retry logic and batching."""
    print("⏳ Fetching live market data...")
    stock_data = {}
    failed_tickers = []
    
    # Process tickers in batches to avoid rate limiting
    for i in range(0, len(tickers), BATCH_SIZE):
        batch = tickers[i:i + BATCH_SIZE]
        print(f"   Processing batch {i//BATCH_SIZE + 1}/{(len(tickers)-1)//BATCH_SIZE + 1}: {', '.join(batch)}")
        
        retry_count = 0
        batch_success = False
        
        while retry_count < MAX_RETRIES and not batch_success:
            try:
                # Fetch data for the batch
                data = yf.download(batch, period='5d', interval='1d', progress=False, threads=False)
                
                # If only one ticker in batch
                if len(batch) == 1:
                    ticker = batch[0]
                    if 'Close' in data and not data.empty:
                        close_series = data['Close']
                        if len(close_series) >= 2:
                            current_price = float(close_series.iloc[-1])
                            prev_close = float(close_series.iloc[-2])
                            stock_data[ticker] = {
                                'current_price_usd': current_price,
                                'prev_close_usd': prev_close,
                            }
                            print(f"      ✓ {ticker}: ${current_price:.2f}")
                        else:
                            print(f"      ✗ {ticker}: Insufficient data")
                            failed_tickers.append(ticker)
                    else:
                        print(f"      ✗ {ticker}: No data returned")
                        failed_tickers.append(ticker)
                else:
                    # Multiple tickers
                    if 'Close' in data.columns and not data.empty:
                        for ticker in batch:
                            try:
                                # Handle MultiIndex columns
                                if isinstance(data.columns, pd.MultiIndex):
                                    close_series = data['Close'][ticker]
                                else:
                                    close_series = data['Close']
                                
                                if isinstance(close_series, pd.Series) and len(close_series) >= 2:
                                    # Get non-NaN values
                                    valid_data = close_series.dropna()
                                    if len(valid_data) >= 2:
                                        current_price = float(valid_data.iloc[-1])
                                        prev_close = float(valid_data.iloc[-2])
                                        stock_data[ticker] = {
                                            'current_price_usd': current_price,
                                            'prev_close_usd': prev_close,
                                        }
                                        print(f"      ✓ {ticker}: ${current_price:.2f}")
                                    else:
                                        print(f"      ✗ {ticker}: Insufficient valid data")
                                        failed_tickers.append(ticker)
                                else:
                                    print(f"      ✗ {ticker}: Invalid data format")
                                    failed_tickers.append(ticker)
                            except Exception as e:
                                print(f"      ✗ {ticker}: Error - {str(e)}")
                                failed_tickers.append(ticker)
                    else:
                        print(f"      ✗ Batch failed: No Close data in response")
                        failed_tickers.extend(batch)
                
                batch_success = True
                
            except Exception as e:
                retry_count += 1
                if retry_count < MAX_RETRIES:
                    print(f"      ⚠️  Batch failed (attempt {retry_count}/{MAX_RETRIES}), retrying in {RETRY_DELAY}s...")
                    time.sleep(RETRY_DELAY)
                else:
                    print(f"      ❌ Batch failed after {MAX_RETRIES} attempts: {str(e)}")
                    failed_tickers.extend(batch)
        
        # Delay between batches to avoid rate limiting
        if i + BATCH_SIZE < len(tickers):
            time.sleep(BATCH_DELAY)
    
    # Retry failed tickers individually
    if failed_tickers:
        print(f"\n🔄 Retrying {len(failed_tickers)} failed tickers individually...")
        for ticker in failed_tickers[:]:  # Create a copy to iterate
            retry_count = 0
            success = False
            
            while retry_count < MAX_RETRIES and not success:
                try:
                    print(f"   Retrying {ticker} (attempt {retry_count + 1}/{MAX_RETRIES})...")
                    ticker_obj = yf.Ticker(ticker)
                    hist = ticker_obj.history(period='5d')
                    
                    if not hist.empty and 'Close' in hist.columns:
                        close_series = hist['Close'].dropna()
                        if len(close_series) >= 2:
                            current_price = float(close_series.iloc[-1])
                            prev_close = float(close_series.iloc[-2])
                            stock_data[ticker] = {
                                'current_price_usd': current_price,
                                'prev_close_usd': prev_close,
                            }
                            print(f"      ✓ {ticker}: ${current_price:.2f}")
                            failed_tickers.remove(ticker)
                            success = True
                        else:
                            retry_count += 1
                    else:
                        retry_count += 1
                    
                    if not success and retry_count < MAX_RETRIES:
                        time.sleep(RETRY_DELAY)
                        
                except Exception as e:
                    retry_count += 1
                    if retry_count < MAX_RETRIES:
                        time.sleep(RETRY_DELAY)
                    else:
                        print(f"      ❌ {ticker}: Failed after {MAX_RETRIES} attempts")
    
    print(f"\n✅ Successfully fetched data for {len(stock_data)}/{len(tickers)} tickers")
    if failed_tickers:
        print(f"❌ Failed to fetch: {', '.join(failed_tickers)}")
    
    return stock_data

def analyze_portfolio(portfolio_df: pd.DataFrame, market_data: Dict[str, Dict[str, float]], exchange_rate: float) -> pd.DataFrame:
    """Calculates all necessary metrics for the summary table."""
    results = []

    for index, row in portfolio_df.iterrows():
        ticker = row['ticker']
        if ticker not in market_data:
            continue

        shares = row['shares_held']
        avg_cost_usd = row['avg_cost_usd']
        
        current_price_usd = market_data[ticker]['current_price_usd']
        prev_close_usd = market_data[ticker]['prev_close_usd']

        daily_change_usd = current_price_usd - prev_close_usd
        daily_percent_change = (daily_change_usd / prev_close_usd) * 100

        current_value_usd = current_price_usd * shares
        current_value_inr = current_value_usd * exchange_rate

        total_investment_usd = avg_cost_usd * shares
        total_investment_inr = total_investment_usd * exchange_rate
        
        total_gain_loss_usd = current_value_usd - total_investment_usd
        total_gain_loss_inr = total_gain_loss_usd * exchange_rate
        
        daily_change_inr = (daily_change_usd * shares) * exchange_rate

        results.append({
            'Ticker': ticker,
            'Shares': shares,
            'Avg Cost (USD)': avg_cost_usd,
            'Current Price (USD)': current_price_usd,
            'Prev Close (USD)': prev_close_usd,
            'Daily Change (%)': daily_percent_change,
            'Daily Change (INR)': daily_change_inr,
            'Investment (USD)': total_investment_usd,
            'Current Value (USD)': current_value_usd,
            'Investment (INR)': total_investment_inr,
            'Current Value (INR)': current_value_inr,
            'Gain/Loss (USD)': total_gain_loss_usd,
            'Gain/Loss (INR)': total_gain_loss_inr,
        })

    return pd.DataFrame(results)

def create_excel_report(summary_df: pd.DataFrame, exchange_rate: float, file_path: str):
    """Creates an Excel file with Summary and Details sheets."""
    
    # Calculate totals
    total_investment_usd = summary_df['Investment (USD)'].sum()
    total_current_value_usd = summary_df['Current Value (USD)'].sum()
    total_gain_loss_usd = summary_df['Gain/Loss (USD)'].sum()
    total_gain_loss_pct = (total_gain_loss_usd / total_investment_usd * 100) if total_investment_usd > 0 else 0
    
    total_investment_inr = summary_df['Investment (INR)'].sum()
    total_current_value_inr = summary_df['Current Value (INR)'].sum()
    total_gain_loss_inr = summary_df['Gain/Loss (INR)'].sum()
    
    # Calculate percentages for each stock
    summary_df['% of Total Value'] = (summary_df['Current Value (INR)'] / total_current_value_inr * 100)
    summary_df['% of Total Investment'] = (summary_df['Investment (INR)'] / total_investment_inr * 100)
    
    # Create workbook
    wb = Workbook()
    
    # --- SHEET 1: SUMMARY ---
    ws_summary = wb.active
    ws_summary.title = "Portfolio Summary"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary['A1'] = "PORTFOLIO SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'] = f"Exchange Rate: 1 USD = ₹{exchange_rate:,.2f}"
    
    # USD Summary
    row = 5
    ws_summary[f'A{row}'] = "USD Summary"
    ws_summary[f'A{row}'].font = title_font
    ws_summary.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ['Metric', 'Value', '', '']
    for col, header in enumerate(headers[:2], 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    usd_data = [
        ['Total Investment', total_investment_usd],
        ['Current Value', total_current_value_usd],
        ['Total Gain/Loss', total_gain_loss_usd],
        ['Return (%)', total_gain_loss_pct]
    ]
    
    for data_row in usd_data:
        row += 1
        ws_summary[f'A{row}'] = data_row[0]
        ws_summary[f'A{row}'].border = border
        cell = ws_summary.cell(row=row, column=2, value=data_row[1])
        cell.border = border
        if data_row[0] == 'Return (%)':
            cell.number_format = '0.00"%"'
        else:
            cell.number_format = '$#,##0.00'
        
        # Color code gain/loss
        if data_row[0] in ['Total Gain/Loss', 'Return (%)']:
            if data_row[1] >= 0:
                cell.font = Font(color="008000", bold=True)
            else:
                cell.font = Font(color="FF0000", bold=True)
    
    # INR Summary
    row += 2
    ws_summary[f'A{row}'] = "INR Summary"
    ws_summary[f'A{row}'].font = title_font
    ws_summary.merge_cells(f'A{row}:D{row}')
    
    row += 1
    for col, header in enumerate(headers[:2], 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    inr_data = [
        ['Total Investment', total_investment_inr],
        ['Current Value', total_current_value_inr],
        ['Total Gain/Loss', total_gain_loss_inr],
        ['Return (%)', total_gain_loss_pct]
    ]
    
    for data_row in inr_data:
        row += 1
        ws_summary[f'A{row}'] = data_row[0]
        ws_summary[f'A{row}'].border = border
        cell = ws_summary.cell(row=row, column=2, value=data_row[1])
        cell.border = border
        if data_row[0] == 'Return (%)':
            cell.number_format = '0.00"%"'
        else:
            cell.number_format = '₹#,##0.00'
        
        if data_row[0] in ['Total Gain/Loss', 'Return (%)']:
            if data_row[1] >= 0:
                cell.font = Font(color="008000", bold=True)
            else:
                cell.font = Font(color="FF0000", bold=True)
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # --- SHEET 2: DETAILS ---
    ws_details = wb.create_sheet("Stock Details")
    
    # Prepare detailed data
    detail_cols = [
        'Ticker', 'Shares', 'Avg Cost (USD)', 'Current Price (USD)', 
        'Daily Change (%)', 'Investment (USD)', 'Current Value (USD)', 
        'Gain/Loss (USD)', 'Investment (INR)', 'Current Value (INR)', 
        'Gain/Loss (INR)', '% of Total Value', '% of Total Investment'
    ]
    
    detail_df = summary_df[detail_cols].sort_values(by='Current Value (INR)', ascending=False)
    
    # Write headers
    for col_idx, col_name in enumerate(detail_cols, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, row_data in enumerate(detail_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Format numbers
            col_name = detail_cols[col_idx - 1]
            if 'USD' in col_name and col_name != 'Avg Cost (USD)' and col_name != 'Current Price (USD)':
                cell.number_format = '$#,##0.00'
            elif col_name in ['Avg Cost (USD)', 'Current Price (USD)']:
                cell.number_format = '$#,##0.00'
            elif 'INR' in col_name:
                cell.number_format = '₹#,##0.00'
            elif '%' in col_name:
                cell.number_format = '0.00"%"'
            elif col_name == 'Shares':
                cell.number_format = '0.00'
            
            # Color code gain/loss
            if 'Gain/Loss' in col_name:
                if value >= 0:
                    cell.font = Font(color="008000")
                else:
                    cell.font = Font(color="FF0000")
    
    # Adjust column widths
    for col_idx, col_name in enumerate(detail_cols, 1):
        ws_details.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Save workbook
    wb.save(file_path)
    print(f"\n✅ Excel report saved to: {file_path}")

def generate_summary_table(summary_df: pd.DataFrame, sort_key_name: str, sort_key_col: str, ascending: bool):
    """Generates and prints the formatted summary table."""
    display_df = summary_df.sort_values(by=sort_key_col, ascending=ascending).copy()
    
    display_cols = [
        'Ticker', 'Shares', 'Current Price (USD)', 'Daily Change (%)', 
        'Current Value (INR)', 'Gain/Loss (INR)', 'Investment (INR)'
    ]
    
    display_df = display_df[display_cols].copy()
    
    # Format for display
    display_df['Current Price (USD)'] = display_df['Current Price (USD)'].apply(lambda x: f"${x:,.2f}")
    for col_name in ['Current Value (INR)', 'Gain/Loss (INR)', 'Investment (INR)']:
        display_df[col_name] = display_df[col_name].apply(lambda x: f"₹{x:,.0f}")
    display_df['Daily Change (%)'] = display_df['Daily Change (%)'].apply(lambda x: f"{x:+.2f}%")
    
    print(f"\n--- Portfolio Summary (Sorted by: {sort_key_name.title()}) ---")
    print(tabulate(display_df, headers='keys', tablefmt='fancy_grid', showindex=False))

def generate_suggestions(summary_df: pd.DataFrame):
    """Provides actionable advice based on portfolio performance."""
    print("\n--- Actionable Trading Suggestions for Today (₹) ---")

    total_daily_change_inr = summary_df['Daily Change (INR)'].sum()
    
    if total_daily_change_inr < -1000:
        print("🟡 Market is volatile today. Wait for consolidation before deploying fresh capital.")
        print("   -> Tip: Identify your highest conviction stocks that are down >2% and prepare funds for a potential buy tomorrow.")
    else:
        print("🟢 Portfolio performance is stable or positive today.")
    
    profit_targets = summary_df[
        (summary_df['Gain/Loss (INR)'] > 50000) & 
        (summary_df['Daily Change (%)'] > 2.0)
    ].sort_values(by='Daily Change (%)', ascending=False)
    
    if not profit_targets.empty:
        print("\n💰 BOOK PROFIT OPPORTUNITIES (Consider trimming positions):")
        for _, row in profit_targets.head(3).iterrows():
            print(f"   -> {row['Ticker']}: Up {row['Daily Change (%)']:+.2f}% today (Total Gain: ₹{row['Gain/Loss (INR)']:,.0f}). Consider selling 5-10% of your shares.")
    
    buy_targets = summary_df[
        (summary_df['Gain/Loss (INR)'] < 0) & 
        (summary_df['Daily Change (%)'] < -1.0)
    ].sort_values(by='Daily Change (%)', ascending=True)

    if not buy_targets.empty:
        print("\n🛒 BUY MORE OPPORTUNITIES (Consider averaging down on losing positions):")
        for _, row in buy_targets.head(3).iterrows():
            print(f"   -> {row['Ticker']}: Down {row['Daily Change (%)']:+.2f}% today (Total Loss: ₹{abs(row['Gain/Loss (INR)']):,.0f}). This might be a good entry point if your long-term thesis is intact.")

def main():
    """Main execution flow."""
    portfolio_df = read_portfolio_data(CSV_FILE_PATH)
    
    tickers = portfolio_df['ticker'].unique().tolist()
    market_data = fetch_stock_data(tickers)
    
    valid_tickers = list(market_data.keys())
    portfolio_df = portfolio_df[portfolio_df['ticker'].isin(valid_tickers)]
    
    if portfolio_df.empty:
        print("Analysis halted as no valid market data was retrieved.")
        return

    summary_df = analyze_portfolio(portfolio_df, market_data, USD_TO_INR)
    
    # Generate Console Outputs
    #generate_summary_table(summary_df, 'Highest Gainer Today', 'Daily Change (%)', False)
    #generate_summary_table(summary_df, 'Highest Value', 'Current Value (INR)', False)
    #generate_summary_table(summary_df, 'Highest Investment', 'Investment (INR)', False)
    generate_suggestions(summary_df)
    
    # Generate Excel Report
    create_excel_report(summary_df, USD_TO_INR, EXCEL_OUTPUT_FILE)

if __name__ == '__main__':
    main()